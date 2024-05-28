import csv

import openpyxl


def read_q_and_a_catalog(path):
    q_and_a_catalog = []
    wb = openpyxl.load_workbook(path)

    sheet = wb.active
    # adjust range according to Excel sheet
    for i in range(2, 21):
        unit = {"nr": i - 1, "question": sheet[f'B{i}'].value, "answer": sheet[f'C{i}'].value}
        q_and_a_catalog.append(unit)

    return q_and_a_catalog


def write_answers_to_text_file(answers_catalog: list[dict]):
    gen_answers_path = "./questions_and_answers/gen_answers.txt"
    ref_answers_path = "./questions_and_answers/ref_answers.txt"

    with open(gen_answers_path, 'w') as gen_answers_file, open(ref_answers_path, 'w') as ref_answers_file:
        for entry in answers_catalog:
            gen_answers_file.write(entry['gen_answer'] + '\n')
            ref_answers_file.write(entry['ref_answer'] + '\n')


def write_answers_to_excel(embedding_model, llm, data):
    # Laden oder Erstellen der Excel-Datei
    try:
        wb = openpyxl.load_workbook("./questions_and_answers/generated_answers.xlsx")
    except FileNotFoundError:
        wb = openpyxl.Workbook()  # Wenn die Datei nicht existiert, erstellen Sie eine neue

    # Wählen oder Erstellen des richtigen Sheets
    if embedding_model == "OpenAI":
        if llm == "gpt-4-turbo":
            sheet_name = "openai-gpt4-with-5-nodes"
        else:
            sheet_name = "openai"
    else:
        sheet_name = "local"

    if sheet_name not in wb.sheetnames:
        sheet = wb.create_sheet(sheet_name)
    else:
        sheet = wb[sheet_name]

    # Finden der nächsten freien Zeile
    next_row = sheet.max_row + 1

    # Wenn das Sheet neu ist oder keine Header hat, erstellen Sie Spaltenüberschriften
    if next_row == 1:
        # Verwenden Sie das erste Dictionary in der Liste, um die Spaltenüberschriften zu definieren
        headers = data[0].keys()
        for col, key in enumerate(headers, start=1):
            sheet.cell(row=1, column=col, value=key)

    # Schreiben der Daten in die Excel-Datei
    for entry in data:
        # print(f"Nr: {next_row}, Frage: {entry['generated_answer']}")
        # Gehen Sie durch die Zeilen und füllen Sie die Zellen aus
        for col, key in enumerate(entry.keys(), start=1):
            if key == "nodes_similarity_score":
                continue
            value = entry.get(key, "")
            sheet.cell(row=next_row, column=col, value=value)
        next_row += 1  # Gehen Sie zur nächsten Zeile über

    # Speichern und Schließen der Arbeitsmappe
    wb.save("./questions_and_answers/generated_answers.xlsx")
    wb.close()


def write_evaluations_to_excel(embedding_model, llm, data):
    # Laden oder Erstellen der Excel-Datei
    try:
        wb = openpyxl.load_workbook("./questions_and_answers/generated_answers.xlsx")
    except FileNotFoundError:
        wb = openpyxl.Workbook()  # Wenn die Datei nicht existiert, erstellen Sie eine neue

    # Wählen oder Erstellen des richtigen Sheets
    if embedding_model == "OpenAI":
        if llm == "gpt-4-turbo":
            sheet_name = "openai_gpt4_evaluations_with_5_nodes"
        else:
            sheet_name = "openai_evaluations"
    else:
        sheet_name = "local_evaluations"

    if sheet_name not in wb.sheetnames:
        sheet = wb.create_sheet(sheet_name)
    else:
        sheet = wb[sheet_name]

    # Finden der nächsten freien Zeile
    next_row = sheet.max_row + 1

    # Wenn das Sheet neu ist oder keine Header hat, erstellen Sie Spaltenüberschriften
    if next_row == 1:
        # Verwenden Sie das erste Dictionary in der Liste, um die Spaltenüberschriften zu definieren
        headers = data[0].keys()
        for col, key in enumerate(headers, start=1):
            sheet.cell(row=1, column=col, value=key)

    # Schreiben der Daten in die Excel-Datei
    for entry in data:
        # Gehen Sie durch die Zeilen und füllen Sie die Zellen aus
        for col, key in enumerate(entry.keys(), start=1):
            sheet.cell(row=next_row, column=col, value=entry[key])
        next_row += 1  # Gehen Sie zur nächsten Zeile über

    # Speichern und Schließen der Arbeitsmappe
    wb.save("./questions_and_answers/generated_answers.xlsx")
    wb.close()
